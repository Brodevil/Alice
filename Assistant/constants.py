import openpyxl
import pprint


def readExel(path):
    """Read the Exel file using openpyxl and return the 2d list as 1st will the time and 2nd will the purpose """

    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    records = list()

    # Loop will returning all columns name
    for j in range(1, sheet_obj.max_row + 1):
        for i in range(1, max_col + 1):
            if i == 2 or i == 3 or i == 5:
                content = sheet_obj.cell(row=j, column=i)
                records.append(content.value)
    records = {records[i]:[records[i+1], records[i+2]] for i in range(0, len(records)-1, 3)}
    return records
         

class Contacts:
    emails = {
        "shivam":"shivamgopale31@gmail.com",
        "aniket":"aniketbhagat0060@gmail.com",
        "anii":"arunagovindbhagat@gmail.com",
        "harsh":"harshkhapare07@gmail.com",
        "darshan":"darshanwarghadre23@gmail.com",
        "aditya":"adityamhaisdhune15@gmail.com",
        "sibam":"shivamgopale31@gmail.com",
        "rohan":"rathodrohan409@gmail.com",
        "archit":"archit.ghadshi@gmail.com",
        "sharvari":"sharvarimolke@gmail.com",
        "saniya":"patilsaniya758@gmail.com",
        "papa":"cbcsudhir@gmail.com",
        "ekta":"ektachoudhary9892@gmail.com",
        "gaurav":"gaurav.y0884@gmail.com",
        "vaishnavi":"kharadevaishnavi608@gmail.com",
        "gautam":"chaudharygautam9963@gmail.com",
        "smit":"smitgondane1919@gmail.com"
    }   



ERROR_REPLIES = [
    "Please don't do that.",
    "You have to stop.",
    "Do you mind?",
    "In the future, don't do that.",
    "That was a mistake.",
    "You blew it.",
    "You're bad at computers.",
    "Are you trying to kill me?",
    "Noooooo!!",
    "I can't believe you've done this",
]

NEGATIVE_REPLIES = [
    "Noooooo!!",
    "Nope.",
    "I'm sorry Dave, I'm afraid I can't do that.",
    "I don't think so.",
    "Not gonna happen.",
    "Out of the question.",
    "Huh? No.",
    "Nah.",
    "Naw.",
    "Not likely.",
    "No way, José.",
    "Not in a million years.",
    "Fat chance.",
    "Certainly not.",
    "NEGATORY.",
    "Nuh-uh.",
    "Not in my house!",
]

POSITIVE_REPLIES = [
    "Yep.",
    "Absolutely!",
    "Can do!",
    "Affirmative!",
    "Yeah okay.",
    "Sure.",
    "Sure thing!",
    "You're the boss!",
    "Okay.",
    "No problem.",
    "I got you.",
    "Alright.",
    "You got it!",
    "ROGER THAT",
    "Of course!",
    "Aye aye, cap'n!",
    "I'll allow it.",
]


if __name__ == "__main__":
    readExel(r"contactinfo.xlsx")