import openpyxl
import os

__all__ = ["contactInfo", "deleteUnwantedFiles", "openApplication", "DailyWorksExel"]


def contactInfo(path):
    """Read the Exel file using openpyxl and return the dictionary containing email id and phone number"""

    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    records = dict()

    # Loop will returning all columns name
    for row in range(3, sheet_obj.max_row - 1):
        name, email, phone_number = (
            sheet_obj.cell(row=row, column=1),
            sheet_obj.cell(row=row, column=2),
            sheet_obj.cell(row=row, column=3),
        )
        records.update({name.value: [email.value, phone_number.value]})
    else:
        return records


def DailyWorksExel(path):
    """Daily task are noted in the file which had been read here and return as a dict
    Time : Task/Work to do
    """
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    tasks = dict()

    for j in range(3, sheet_obj.max_row + 1):
        workTime = sheet_obj.cell(row=j, column=1)
        work = sheet_obj.cell(row=j, column=2)
        tasks.update({workTime.value: work.value})
    return tasks


def deleteUnwantedFiles():
    """
    The following directory contain the temporary files and just unwanted files
    """
    unwantedFiles = [
        r"C:\Windows\Temp",
        r"C:\Users\ADMIN\AppData\Local\Temp",
        r"C:\Windows\Prefetch",
    ]
    for f in unwantedFiles:
        for file in os.listdir(f):
            try:
                os.remove(os.path.join(f, file))
            except PermissionError:
                pass


def openApplication(ApplicationName: str, installedApplicationPath: str):
    """To match to queary with the available application in the Application folder i.e.
    the .lnk files And opening or launching the most matching queary name of applications"""

    installed_application_shortcut_path = os.listdir(installedApplicationPath)

    for app in installed_application_shortcut_path:
        for name in ApplicationName.split():
            if name.lower() in app[:-4].lower().split():
                os.startfile(os.path.join(installedApplicationPath, app))
                return app.split(".")[0]
    return None
