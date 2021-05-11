import openpyxl
import pprint


def contactInfo(path):
    """Read the Exel file using openpyxl and return the dictionary containing email id and phone number """

    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    max_col = sheet_obj.max_column
    records = list()

    # Loop will returning all columns name
    for j in range(1, sheet_obj.max_row + 1):
        for i in range(1, max_col + 1):
            if i == 2 or i == 3 or i == 5:
                content = sheet_obj.cell(row=j, column=i)
                records.append(content.value)
    records = {records[i]:[records[i+1], records[i+2]] for i in range(0, len(records)-1, 3)}
    return records


if __name__ == "__main__":
    pprint.pprint(contactInfo(r"M:\ADMIN\Critical Data\VS-Code\Alice\contactinfo.xlsx"))